import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta
import json

random.seed(42)
np.random.seed(42)

# === GENERATE HELPDESK TICKET DATA ===
categories = ["Network/VPN", "Hardware", "Software", "Account Access", "Email", "Onboarding", "Security"]
agents = ["Ernestina Z.", "James K.", "Ama S.", "Kofi M."]
priorities = ["P1 - Critical", "P2 - High", "P3 - Medium", "P4 - Low"]
priority_weights = [0.05, 0.20, 0.50, 0.25]
statuses = ["Resolved", "Resolved", "Resolved", "Closed", "In Progress", "Pending"]

sla_targets = {"P1 - Critical": 1, "P2 - High": 4, "P3 - Medium": 8, "P4 - Low": 24}

tickets = []
for i in range(320):
    ticket_id = f"TKT-{2000+i:04d}"
    cat = random.choice(categories)
    priority = np.random.choice(priorities, p=priority_weights)
    created = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 364), hours=random.randint(7, 18))
    agent = random.choice(agents)
    sla_target = sla_targets[priority]

    # Resolution time: varies by priority
    if priority == "P1 - Critical":
        res_hours = max(0.5, np.random.exponential(0.8))
    elif priority == "P2 - High":
        res_hours = max(0.5, np.random.exponential(3.0))
    elif priority == "P3 - Medium":
        res_hours = max(0.5, np.random.exponential(6.5))
    else:
        res_hours = max(0.5, np.random.exponential(18))

    met_sla = res_hours <= sla_target
    fcr = random.random() < 0.87  # 87% FCR

    status = "Resolved" if random.random() < 0.93 else "Closed" if random.random() < 0.5 else "In Progress"
    csat = round(random.uniform(3.5, 5.0), 1) if status in ["Resolved", "Closed"] else None

    tickets.append({
        "ticket_id": ticket_id,
        "created_date": created.strftime("%Y-%m-%d"),
        "month": created.strftime("%b %Y"),
        "category": cat,
        "priority": priority,
        "assigned_agent": agent,
        "status": status,
        "resolution_hours": round(res_hours, 1),
        "sla_target_hours": sla_target,
        "met_sla": "Yes" if met_sla else "No",
        "first_contact_resolved": "Yes" if fcr else "No",
        "csat_score": csat,
    })

df = pd.DataFrame(tickets)
df.to_csv("/home/claude/project1/data/helpdesk_tickets.csv", index=False)

# === BUILD HELPDESK EXCEL DASHBOARD ===
NAVY = "1B3A5C"; TEAL = "1D7A6B"; WHITE = "FFFFFF"; LTBLUE = "EAF2FA"
hf = Font(name="Arial", size=11, bold=True, color=WHITE)
hfill = PatternFill("solid", fgColor=NAVY)
nf = Font(name="Arial", size=10)
bf = Font(name="Arial", size=10, bold=True)
lf = Font(name="Arial", size=11, bold=True, color=NAVY)
kf = Font(name="Arial", size=20, bold=True, color=TEAL)
ctr = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# --- SHEET 1: KPI DASHBOARD ---
ws = wb.active
ws.title = "KPI Dashboard"

# KPIs
total_tickets = len(df)
resolved = df[df["status"].isin(["Resolved", "Closed"])]
sla_met = round(len(df[df["met_sla"] == "Yes"]) / len(df) * 100, 1)
fcr_rate = round(len(df[df["first_contact_resolved"] == "Yes"]) / len(df) * 100, 1)
avg_res = round(df["resolution_hours"].mean(), 1)
avg_csat = round(df["csat_score"].dropna().mean(), 2)

ws["A1"] = "IT Helpdesk KPI Dashboard"
ws["A1"].font = Font(name="Arial", size=16, bold=True, color=NAVY)
ws["A2"] = "Period: Jan–Dec 2024  |  Agent: Ernestina Thembi Zah  |  320 tickets"
ws["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

kpis = [
    ("Total Tickets", str(total_tickets)),
    ("SLA Compliance", f"{sla_met}%"),
    ("FCR Rate", f"{fcr_rate}%"),
    ("Avg Resolution", f"{avg_res}h"),
    ("Avg CSAT", f"{avg_csat}/5.0"),
]
for j, (label, val) in enumerate(kpis, 1):
    ws.cell(row=4, column=j, value=label).font = Font(name="Arial", size=10, color="6B7280")
    ws.cell(row=4, column=j).alignment = ctr
    ws.cell(row=4, column=j).fill = PatternFill("solid", fgColor=LTBLUE)
    ws.cell(row=5, column=j, value=val).font = kf
    ws.cell(row=5, column=j).alignment = ctr
    ws.cell(row=5, column=j).fill = PatternFill("solid", fgColor=LTBLUE)
    ws.column_dimensions[get_column_letter(j)].width = 18
ws.row_dimensions[5].height = 30

# Tickets by category table
ws["A8"] = "Tickets by Category"
ws["A8"].font = lf
cat_data = df.groupby("category").agg(total=("ticket_id","count"), sla=("met_sla", lambda x: (x=="Yes").sum()), fcr=("first_contact_resolved", lambda x: (x=="Yes").sum())).reset_index()
headers = ["Category", "Total Tickets", "SLA Met", "SLA %", "FCR Count", "FCR %"]
for j, h in enumerate(headers, 1):
    c = ws.cell(row=9, column=j, value=h); c.font = hf; c.fill = hfill
for i, row in cat_data.iterrows():
    r = 10 + i
    ws.cell(row=r, column=1, value=row["category"]).font = nf
    ws.cell(row=r, column=2, value=int(row["total"])).font = nf
    ws.cell(row=r, column=3, value=int(row["sla"])).font = nf
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = "0%"
    ws.cell(row=r, column=5, value=int(row["fcr"])).font = nf
    ws.cell(row=r, column=6, value=f"=E{r}/B{r}").number_format = "0%"

# Category chart
chart1 = BarChart()
chart1.title = "Tickets by Category"
chart1.style = 10
data = Reference(ws, min_col=2, min_row=9, max_row=9+len(cat_data))
cats = Reference(ws, min_col=1, min_row=10, max_row=9+len(cat_data))
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 8; chart1.width = 12
ws.add_chart(chart1, "H8")

# Priority breakdown table
start = 10 + len(cat_data) + 3
ws.cell(row=start-1, column=1, value="Tickets by Priority").font = lf
pri_data = df.groupby("priority").agg(total=("ticket_id","count"), sla=("met_sla", lambda x: (x=="Yes").sum())).reset_index()
ws.cell(row=start, column=1, value="Priority").font = hf; ws.cell(row=start, column=1).fill = hfill
ws.cell(row=start, column=2, value="Total").font = hf; ws.cell(row=start, column=2).fill = hfill
ws.cell(row=start, column=3, value="SLA Met").font = hf; ws.cell(row=start, column=3).fill = hfill
ws.cell(row=start, column=4, value="SLA %").font = hf; ws.cell(row=start, column=4).fill = hfill
for i, row in pri_data.iterrows():
    r = start + 1 + i
    ws.cell(row=r, column=1, value=row["priority"]).font = nf
    ws.cell(row=r, column=2, value=int(row["total"])).font = nf
    ws.cell(row=r, column=3, value=int(row["sla"])).font = nf
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = "0%"

# --- SHEET 2: TICKET LOG ---
ws2 = wb.create_sheet("Ticket Log")
from openpyxl.utils.dataframe import dataframe_to_rows
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = hf; cell.fill = hfill
ws2.freeze_panes = "A2"
for c in range(1, len(df.columns)+1):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# --- SHEET 3: AGENT PERFORMANCE ---
ws3 = wb.create_sheet("Agent Performance")
ws3["A1"].value = "Agent Performance Summary"
ws3["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY)
agent_data = df.groupby("assigned_agent").agg(
    total=("ticket_id","count"),
    sla_met=("met_sla", lambda x: (x=="Yes").sum()),
    fcr=("first_contact_resolved", lambda x: (x=="Yes").sum()),
    avg_res=("resolution_hours", "mean"),
    avg_csat=("csat_score", "mean")
).reset_index()
headers2 = ["Agent", "Total Tickets", "SLA Met", "SLA %", "FCR Count", "FCR %", "Avg Res (h)", "Avg CSAT"]
for j, h in enumerate(headers2, 1):
    c = ws3.cell(row=3, column=j, value=h); c.font = hf; c.fill = hfill
for i, row in agent_data.iterrows():
    r = 4 + i
    ws3.cell(row=r, column=1, value=row["assigned_agent"]).font = nf
    ws3.cell(row=r, column=2, value=int(row["total"])).font = nf
    ws3.cell(row=r, column=3, value=int(row["sla_met"])).font = nf
    ws3.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = "0%"
    ws3.cell(row=r, column=5, value=int(row["fcr"])).font = nf
    ws3.cell(row=r, column=6, value=f"=E{r}/B{r}").number_format = "0%"
    ws3.cell(row=r, column=7, value=round(row["avg_res"],1)).font = nf
    ws3.cell(row=r, column=8, value=round(row["avg_csat"],2)).font = nf
for c in range(1, 9):
    ws3.column_dimensions[get_column_letter(c)].width = 17

wb.save("/home/claude/project1/dashboard/helpdesk_dashboard.xlsx")
print(f"Helpdesk dashboard saved. {len(df)} tickets | SLA: {sla_met}% | FCR: {fcr_rate}% | CSAT: {avg_csat}")
